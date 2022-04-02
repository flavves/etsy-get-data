# -*- coding: utf-8 -*-
"""
Created on Sun Mar 20 01:48:15 2022

@author: okmen
"""

from bs4 import BeautifulSoup

import urllib.request as urllib2

import openpyxl

import pandas as pd

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
from  selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#link yarat

#burada bütün kategorileri çektim yanlışlıkla

c= urllib2.urlopen("https://www.etsy.com/search?q=car+mats+&s2qii=7&s2qit=as&prq=car+mats")
contents=c.read()
soup=BeautifulSoup(contents)
soup=str(soup)

link_deneme=soup.split('https://www.etsy.com/listing/')
link_deneme.pop(0)

link_deneme[1][:10]


butun_linkler=[]


for i in link_deneme:

    link_aldim="https://www.etsy.com/listing/"+i[:10]
    butun_linkler.append(link_aldim)



driver = webdriver.Chrome(executable_path=r'chromedriver.exe')

driver.get(butun_linkler[3])

html = driver.page_source
soup = BeautifulSoup(html) 



variant_1_uzunluk=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]').text
variant_1_uzunluk=variant_1_uzunluk.split("\n")
variant_1_uzunluk.pop(0)
variant_1_uzunluk.pop(-1)
variant_1_uzunluk=len(variant_1_uzunluk)

varyasyon1_sayac=2
varyasyon2_uzunluk=2

varyasyon2_var_mi=False


kitap = openpyxl.load_workbook("deneme.xlsx")
sayfa = kitap.get_sheet_by_name("Sayfa1")

sira=2

varyasyon_sayisi=0
sayac_gosterme=0
for k in butun_linkler:
    try:
            
        print(str(len(butun_linkler))+"/"+str(sayac_gosterme))
        sayac_gosterme=sayac_gosterme+1
        varyasyon_sayisi=0
        driver.get(k)   
        
        
        kod=k.split("/")[-1]
        
        #varyasyon sayısının öğrenilmesi gerekiyor
        
        try:                                                                    
            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div/div/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p'))
            )
        finally:
            pass
        
        
        
        try:                                                                    
            element = WebDriverWait(driver, 0.5).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="inventory-variation-select-0"]'))
            )
            varyasyon_sayisi=varyasyon_sayisi+1
            
            variant_1_uzunluk=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]').text
            variant_1_uzunluk=variant_1_uzunluk.split("\n")
            variant_1_uzunluk.pop(0)
            variant_1_uzunluk.pop(-1)
            variant_1_uzunluk=len(variant_1_uzunluk)
            
        except:
            pass
        
        try:                                                                    
            element = WebDriverWait(driver, 0.5).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="inventory-variation-select-1"]'))
            )
            varyasyon_sayisi=varyasyon_sayisi+1
            
            varyasyon2_uzunluk=driver.find_element_by_xpath('//*[@id="inventory-variation-select-1"]').text
            varyasyon2_uzunluk=varyasyon2_uzunluk.split("\n")
            varyasyon2_uzunluk.pop(0)
            varyasyon2_uzunluk.pop(-1)
            varyasyon2_uzunluk=len(varyasyon2_uzunluk)
            
        except:
            pass
        
        
    
        
        #varyasyon olmadan
        
        if varyasyon_sayisi==0:
            adi=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[2]/h1').text
            try:
                fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text
                fiyat=fiyat.split("\n")[1]
            except:
                fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text

            
            sayfa.cell(row=sira,column=1,value=adi)
            sayfa.cell(row=sira,column=2,value=kod)
            sayfa.cell(row=sira,column=7,value=fiyat)  
            sira=sira+1
            
        
       
        elif varyasyon_sayisi==1:
            
            
            #tek varyasyonlu okey
            
            
            for varyasyon1_sayac in range(2,variant_1_uzunluk+2):
                
                try:                                                                    
                    element = WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div/div/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p'))
                    )
                finally:
                    pass
                   
                
                try:
                    try:
                        element = WebDriverWait(driver, 0.5).until(
                            EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div/div/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p'))
                        )
                    finally:
                        pass
                 
                
                    driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').click()
                    variant_1_Ad=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').text
                    variant_1_Kod=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').get_attribute('value')
        
                except:
                    pass
                
                adi=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[2]/h1').text
                try:
                    fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text
                    fiyat=fiyat.split("\n")[1]
                except:
                    fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text

                
                sayfa.cell(row=sira,column=1,value=adi)
                sayfa.cell(row=sira,column=2,value=kod)
                sayfa.cell(row=sira,column=3,value=variant_1_Kod)
                #sayfa.cell(row=sira,column=4,value=variant2)
                sayfa.cell(row=sira,column=5,value=variant_1_Ad)
                #sayfa.cell(row=sira,column=6,value=variant2ad)
                sayfa.cell(row=sira,column=7,value=fiyat)  
                sira=sira+1
                
                
        #2 varyasyonlu
        
        elif varyasyon_sayisi==2:
            
            
            
            
            
            for varyasyon1_sayac in range(2,variant_1_uzunluk+2):
                
                
                
                for varyasyon2_sayac in range(2,varyasyon2_uzunluk+2):
                    
                    
                    
                    try:                                                                    
                        element = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div/div/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p'))
                        )
                    finally:
                        pass
                       
                    
                    
                    
                    
                    
                    try:
                        try:
                            element = WebDriverWait(driver, 0.5).until(
                                EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div/div/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p'))
                            )
                        finally:
                            pass
                     
                    except:
                        pass
                    
                    
                    
                    try:
                        
                        driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').click()
                        variant_1_Ad=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').text
                        variant_1_Kod=driver.find_element_by_xpath('//*[@id="inventory-variation-select-0"]/option['+str(varyasyon1_sayac)+']').get_attribute('value')
                        
                        driver.find_element_by_xpath('//*[@id="inventory-variation-select-1"]/option['+str(varyasyon2_sayac)+']').click()
                        variant_2_Ad=driver.find_element_by_xpath('//*[@id="inventory-variation-select-1"]/option['+str(varyasyon2_sayac)+']').text
                        variant_2_Kod=driver.find_element_by_xpath('//*[@id="inventory-variation-select-1"]/option['+str(varyasyon2_sayac)+']').get_attribute('value')
                    
                    except:
                        pass
                    
                    
                    
                    
                    
                    adi=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[2]/h1').text
                    
                    try:
                        fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text
                        fiyat=fiyat.split("\n")[1]
                    except:
                        fiyat=driver.find_element_by_xpath('//*[@id="listing-page-cart"]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/p').text

                        
                    
                    sayfa.cell(row=sira,column=1,value=adi)
                    sayfa.cell(row=sira,column=2,value=kod)
                    sayfa.cell(row=sira,column=3,value=variant_1_Kod)
                    sayfa.cell(row=sira,column=4,value=variant_2_Kod)
                    sayfa.cell(row=sira,column=5,value=variant_1_Ad)
                    sayfa.cell(row=sira,column=6,value=variant_2_Ad)
                    sayfa.cell(row=sira,column=7,value=fiyat)  
                    sira=sira+1
            
    except Exception as e:
        print(e)              
        
        

        
           
        

kitap.save("deneme.xlsx")
kitap.close()






















